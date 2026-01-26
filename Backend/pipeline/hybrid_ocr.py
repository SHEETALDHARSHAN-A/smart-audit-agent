"""
Hybrid OCR Engine Router
Intelligently selects the best OCR engine based on document type and content
"""
import os
import logging
import time
import warnings

from typing import Dict, List, Any, Tuple
import numpy as np

class HybridOCREngine:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.engines = {}
        self._initialize_engines()
    
    def _initialize_engines(self):
        """Lazy-load OCR engines as needed"""
        try:
            from paddleocr import PaddleOCR
            self.engines['paddle'] = PaddleOCR(use_angle_cls=True, lang='en')
            self.logger.info("PaddleOCR initialized")
        except Exception as e:
            self.logger.error(f"Failed to initialize PaddleOCR: {e}")
    
    def _get_tesseract(self):
        """Lazy load Tesseract"""
        if 'tesseract' not in self.engines:
            try:
                import pytesseract
                from PIL import Image
                self.engines['tesseract'] = pytesseract
                self.engines['PIL'] = Image
                self.logger.info("Tesseract initialized")
            except Exception as e:
                self.logger.error(f"Failed to initialize Tesseract: {e}")
                return None
        return self.engines.get('tesseract')
    

    
    def detect_handwriting(self, image_path: str) -> bool:
        """
        Detect if image contains handwriting using heuristics.
        Returns True if handwriting is detected.
        """
        try:
            import cv2
            
            img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
            if img is None:
                return False
            
            # Apply edge detection
            edges = cv2.Canny(img, 50, 150)
            
            # Calculate edge density
            edge_density = np.sum(edges > 0) / edges.size
            
            # Apply line detection
            lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=50, 
                                    minLineLength=50, maxLineGap=10)
            
            # Handwriting indicators:
            # 1. Lower edge density (handwriting is more sparse)
            # 2. Fewer straight lines (handwriting is curved/irregular)
            # 3. High variance in stroke width
            
            num_lines = len(lines) if lines is not None else 0
            
            # Heuristic: if few straight lines and moderate edge density, likely handwriting
            is_handwritten = (num_lines < 20 and 0.02 < edge_density < 0.15)
            
            if is_handwritten:
                self.logger.info(f"Handwriting detected in {image_path}")
            
            return is_handwritten
            
        except Exception as e:
            self.logger.error(f"Handwriting detection failed: {e}")
            return False
    

    
    def detect_document_type(self, file_path: str) -> str:
        """Detect document type from file extension"""
        ext = os.path.splitext(file_path)[1].lower()
        type_map = {
            '.pdf': 'pdf',
            '.png': 'image',
            '.jpg': 'image',
            '.jpeg': 'image',
            '.tiff': 'image',
            '.bmp': 'image',
            '.docx': 'docx',
            '.doc': 'doc',
            '.xlsx': 'xlsx',
            '.xls': 'xls'
        }
        return type_map.get(ext, 'unknown')
    
    def extract_with_tesseract(self, image_path: str) -> str:
        """Extract text using Tesseract OCR"""
        tesseract = self._get_tesseract()
        if not tesseract:
            return ""
        
        try:
            from PIL import Image
            img = Image.open(image_path)
            text = tesseract.image_to_string(img, config='--psm 6')
            self.logger.info(f"Tesseract extracted {len(text)} characters")
            return text
        except Exception as e:
            self.logger.error(f"Tesseract extraction failed: {e}")
            return ""
    

    
    def extract_tables(self, pdf_path: str) -> List[Dict]:
        """Extract tables from PDF using Camelot"""
        try:
            import camelot
            tables = camelot.read_pdf(pdf_path, pages='all', flavor='lattice')
            
            result = []
            for i, table in enumerate(tables):
                result.append({
                    'table_number': i + 1,
                    'data': table.df.to_dict('records'),
                    'accuracy': table.accuracy
                })
            
            self.logger.info(f"Extracted {len(tables)} tables from PDF")
            return result
        except Exception as e:
            self.logger.error(f"Table extraction failed: {e}")
            return []
    
    def extract_from_docx(self, docx_path: str) -> str:
        """Extract text from Word document"""
        try:
            from docx import Document
            doc = Document(docx_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
            self.logger.info(f"Extracted {len(text)} characters from DOCX")
            return text
        except Exception as e:
            self.logger.error(f"DOCX extraction failed: {e}")
            return ""
    
    def extract_from_xlsx(self, xlsx_path: str) -> Dict:
        """Extract data from Excel file"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path)
            result = {}
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
                result[sheet_name] = data
            
            self.logger.info(f"Extracted {len(result)} sheets from XLSX")
            return result
        except Exception as e:
            self.logger.error(f"XLSX extraction failed: {e}")
            return {}
    
    def extract_text_from_pdf(self, pdf_path: str) -> Tuple[str, bool]:
        """Extract text from PDF, detect if scanned"""
        try:
            import pdfplumber
            text = ""
            is_scanned = True
            
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text and page_text.strip():
                        text += page_text + "\n"
                        is_scanned = False
            
            if is_scanned or len(text.strip()) < 50:
                self.logger.info("PDF appears to be scanned, will use OCR")
                return text, True
            else:
                self.logger.info(f"Extracted {len(text)} characters from text-based PDF")
                return text, False
                
        except Exception as e:
            self.logger.error(f"PDF text extraction failed: {e}")
            return "", True
    
    def get_text_blocks(self, file_path: str, use_hybrid: bool = True) -> List[Dict]:
        """
        Main method to extract text using appropriate engine(s)
        """
        doc_type = self.detect_document_type(file_path)
        self.logger.info(f"Processing {doc_type} document: {file_path}")
        
        text_blocks = []
        combined_text = ""
        
        # Handle different document types
        if doc_type == 'docx':
            combined_text = self.extract_from_docx(file_path)
            text_blocks = [{"text": combined_text, "confidence": 1.0, "box": []}]
            
        elif doc_type == 'xlsx':
            data = self.extract_from_xlsx(file_path)
            # Convert Excel data to text representation
            for sheet, rows in data.items():
                combined_text += f"\nSheet: {sheet}\n"
                for row in rows:
                    combined_text += " | ".join([str(cell) for cell in row if cell]) + "\n"
            text_blocks = [{"text": combined_text, "confidence": 1.0, "box": []}]
            
        elif doc_type == 'pdf':
            # Try text extraction first
            pdf_text, is_scanned = self.extract_text_from_pdf(file_path)
            
            if not is_scanned and pdf_text:
                # Text-based PDF
                text_blocks = [{"text": pdf_text, "confidence": 1.0, "box": []}]
            else:
                # Scanned PDF - convert pages to images and preprocess
                text_blocks = self._process_scanned_pdf(file_path)
        
        elif doc_type == 'image':
            # Use hybrid approach for images
            if use_hybrid:
                # Try PaddleOCR first (best for layout)
                if 'paddle' in self.engines:
                    result = self.engines['paddle'].ocr(file_path)
                    if result and result[0]:
                        for line in result[0]:
                            text_blocks.append({
                                "text": line[1][0],
                                "confidence": line[1][1],
                                "box": line[0]
                            })
                
                # Also try Tesseract for comparison
                tesseract_text = self.extract_with_tesseract(file_path)
                if tesseract_text and len(tesseract_text) > len(combined_text):
                    # If Tesseract found more text, use it
                    text_blocks.append({
                        "text": tesseract_text,
                        "confidence": 0.8,
                        "box": [],
                        "engine": "tesseract"
                    })
            else:
                # Default to PaddleOCR
                if 'paddle' in self.engines:
                    result = self.engines['paddle'].ocr(file_path)
                    if result and result[0]:
                        for line in result[0]:
                            text_blocks.append({
                                "text": line[1][0],
                                "confidence": line[1][1],
                                "box": line[0]
                            })
        
        return text_blocks if text_blocks else [{"text": "", "confidence": 0, "box": []}]
    
    def _process_scanned_pdf(self, pdf_path: str) -> List[Dict]:
        """Convert PDF pages to images and run appropriate OCR"""
        text_blocks = []
        
        try:
            import fitz  # PyMuPDF
            import os
            
            pdf_doc = fitz.open(pdf_path)
            temp_dir = "temp_processed"
            os.makedirs(temp_dir, exist_ok=True)
            
            for page_num, page in enumerate(pdf_doc):
                # Convert page to high-res image
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                img_path = os.path.join(temp_dir, f"pdf_page_{page_num}.png")
                pix.save(img_path)
                
                # Use PaddleOCR to get text boxes
                if 'paddle' in self.engines:
                    result = self.engines['paddle'].ocr(img_path)
                    if result and result[0]:
                        for line in result[0]:
                            box = line[0]
                            text = line[1][0]
                            conf = line[1][1]
                            
                            text_blocks.append({
                                "text": text,
                                "confidence": conf,
                                "box": box,
                                "page": page_num + 1,
                                "engine": "paddle"
                            })
                
                self.logger.info(f"Processed PDF page {page_num + 1}")
            
            pdf_doc.close()
            
        except Exception as e:
            self.logger.error(f"PDF preprocessing failed: {e}")
            # Fallback to direct OCR if possible
            if 'paddle' in self.engines:
                result = self.engines['paddle'].ocr(pdf_path)
                if result and result[0]:
                    for line in result[0]:
                        text_blocks.append({
                            "text": line[1][0],
                            "confidence": line[1][1],
                            "box": line[0]
                        })
        
        return text_blocks
    
    def _is_region_handwritten(self, region) -> bool:
        """Detect if a small text region is handwritten"""
        try:
            import cv2
            
            if region is None or region.size == 0:
                return False
            
            # Convert to grayscale
            if len(region.shape) == 3:
                gray = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
            else:
                gray = region
            
            # Handwriting detection heuristics for small regions:
            
            # 1. Edge irregularity - handwriting has more irregular edges
            edges = cv2.Canny(gray, 50, 150)
            edge_density = np.sum(edges > 0) / edges.size
            
            # 2. Stroke width variation - handwriting has more variation
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
            
            # Count horizontal runs to detect stroke consistency
            stroke_widths = []
            for row in binary:
                in_stroke = False
                width = 0
                for pixel in row:
                    if pixel > 0:
                        if not in_stroke:
                            in_stroke = True
                        width += 1
                    else:
                        if in_stroke:
                            stroke_widths.append(width)
                            width = 0
                            in_stroke = False
            
            if len(stroke_widths) > 5:
                stroke_std = np.std(stroke_widths)
                stroke_mean = np.mean(stroke_widths)
                stroke_variation = stroke_std / max(stroke_mean, 1)
            else:
                stroke_variation = 0
            
            # 3. Line straightness - handwriting has fewer straight horizontal lines
            lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=20, 
                                    minLineLength=max(10, region.shape[1]//4), maxLineGap=5)
            num_straight_lines = len(lines) if lines is not None else 0
            
            # Heuristics for handwriting (Strict Mode):
            # - Moderate edge density (not too clean, not too noisy)
            # - High stroke width variation (handwriting has varying pen pressure/width)
            # - Few straight lines (printed text has many straight lines)
            is_handwritten = (
                (0.02 < edge_density < 0.12) and  # Tighter upper bound (printed text is dense)
                (stroke_variation > 0.6 or num_straight_lines < 2)  # Higher variation required
            )
            return is_handwritten
            
        except Exception as e:
            self.logger.debug(f"Region handwriting detection failed: {e}")
            return False
    
    def _preprocess_image(self, image_path: str) -> str:
        """Apply image preprocessing for better OCR accuracy"""
        try:
            import cv2
            import os
            
            # Read image in grayscale
            img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
            if img is None:
                return image_path
            
            # Upscale if small
            height, width = img.shape
            if height < 1000 or width < 1000:
                img = cv2.resize(img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
            
            # Denoise
            img = cv2.GaussianBlur(img, (3, 3), 0)
            
            # Adaptive thresholding for better handling of varying backgrounds
            img = cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                        cv2.THRESH_BINARY, 11, 2)
            
            # Morphological Closing: Connects broken strokes (e.g., fixing "4" misread as "1")
            # Kernel size (2,2) or (3,3) depends on resolution, starting small to avoid merging distinct chars
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
            img = cv2.morphologyEx(img, cv2.MORPH_CLOSE, kernel)
            
            # Save preprocessed image
            temp_dir = "temp_processed"
            os.makedirs(temp_dir, exist_ok=True)
            filename = os.path.basename(image_path)
            save_path = os.path.join(temp_dir, f"prep_{filename}")
            cv2.imwrite(save_path, img)
            
            self.logger.info(f"Image preprocessed: {save_path}")
            return save_path
            
        except Exception as e:
            self.logger.error(f"Image preprocessing failed: {e}")
            return image_path




